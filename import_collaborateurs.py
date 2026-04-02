"""
Script d'import des collaborateurs DFC depuis le fichier Excel
Liste_collaborateurs_DFC_MISE_0_JOUR_LE_2502206.xlsx

Ce script crée :
  1. Les sous-directions manquantes
  2. Les sections pour chaque SD
  3. Un compte utilisateur par collaborateur

Placer ce fichier ET le fichier Excel dans le dossier app\

Usage : python import_collaborateurs.py
"""
import os, sys, re, unicodedata

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django
django.setup()

import openpyxl
from apps.organisation.models import SousDirection, Section, UtilisateurSection
from apps.authentication.models import Utilisateur, Role

def log(msg):  print(f"  ✓  {msg}")
def warn(msg): print(f"  ⚠  {msg}")
def head(msg): print(f"\n{'─'*60}\n  {msg}\n{'─'*60}")

MOT_DE_PASSE = "cie2026"

# ── Mapping catégorie Excel → rôle Django ─────────────────────────────────────
CAT_ROLE = {
    'D':  'dfc',       # Directeur Financier
    'DA': 'da',        # Directeur Adjoint
    'SD': 'sd',        # Sous-Directeur
    'CS': 'chef',      # Chef de Service
    'C':  'cadre',     # Cadre
    'M':  'maitrise',  # Maîtrise
    'EO': 'visiteur',  # Employé Ordinaire
}

# ── Mapping SD Excel → SD en base (get_or_create) ─────────────────────────────
SD_DEFINITIONS = {
    'SDCC':   ('Sous-Direction Comptabilité Clientèle',           '#003F7F'),
    'SDPCC':  ('Sous-Direction Projet et Contrôle Clientèle',     '#0056A8'),
    'SDCG':   ('Sous-Direction Comptabilité Générale',            '#6f42c1'),
    'SDCF':   ('Sous-Direction Comptabilité Fournisseurs',        '#dc3545'),
    'SDTRAC': ('Sous-Direction Trésorerie et Relation AC',        '#17a2b8'),
    'SDRC':   ('Sous-Direction Recouvrement Contentieux',         '#fd7e14'),
    'SDRCCC': ('Sous-Direction Recouvrement CCC',                 '#F5A623'),
    'SDF':    ('Sous-Direction Fiscalité',                        '#1a7cc1'),
    'STAFF':  ('Staff Direction',                                 '#6c757d'),
}

SD_SECTIONS = {
    'SDCC':   ('SC-CC',    'Section Comptabilité Clientèle'),
    'SDPCC':  ('SC-PCC',   'Section Projet et Contrôle'),
    'SDCG':   ('SC-CG',    'Section Comptabilité Générale'),
    'SDCF':   ('SC-CF',    'Section Comptabilité Fournisseurs'),
    'SDTRAC': ('SC-TRAC',  'Section Trésorerie et RAC'),
    'SDRC':   ('SC-RC',    'Section Recouvrement'),
    'SDRCCC': ('SC-RCCC',  'Section Recouvrement CCC'),
    'SDF':    ('SC-FISC',  'Section Fiscalité'),
    'STAFF':  ('SC-STAFF', 'Section Staff Direction'),
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def nettoyer(texte):
    """Supprime les espaces et tabulations parasites."""
    if not texte:
        return ''
    return str(texte).strip().replace('\t', '').replace('\xa0', ' ')

def generer_username(prenom, nom, index):
    """Génère un username unique : prenom.nom (sans accents, sans espaces)."""
    def sans_accent(s):
        return ''.join(
            c for c in unicodedata.normalize('NFD', s)
            if unicodedata.category(c) != 'Mn'
        )
    p = re.sub(r'[^a-z]', '', sans_accent(prenom.split()[0]).lower()) if prenom else 'user'
    n = re.sub(r'[^a-z]', '', sans_accent(nom.split()[0]).lower())    if nom    else str(index)
    base = f"{p}.{n}"
    username = base
    i = 2
    while Utilisateur.objects.filter(username=username).exists():
        username = f"{base}{i}"
        i += 1
    return username


# ── Chargement des rôles ──────────────────────────────────────────────────────
head("Chargement des rôles")
roles = {r.code: r for r in Role.objects.all()}
print(f"  Rôles disponibles : {list(roles.keys())}")
if not roles:
    print("ERREUR : Aucun rôle en base. Lancez : python manage.py loaddata fixtures/roles.json")
    sys.exit(1)


# ── Sous-directions ───────────────────────────────────────────────────────────
head("Sous-directions")
sds = {}
for i, (code, (libelle, couleur)) in enumerate(SD_DEFINITIONS.items(), 1):
    sd, created = SousDirection.objects.get_or_create(
        code=code,
        defaults={
            'libelle':     libelle,
            'description': libelle,
            'mission':     '',
            'couleur':     couleur,
            'actif':       True,
            'ordre':       i,
        }
    )
    sds[code] = sd
    log(f"{'Créée' if created else 'Existe'} : {code} — {libelle}")


# ── Sections ──────────────────────────────────────────────────────────────────
head("Sections")
sections = {}
for code_sd, (code_sec, libelle_sec) in SD_SECTIONS.items():
    sd = sds[code_sd]
    section, created = Section.objects.get_or_create(
        code=code_sec,
        sous_direction=sd,
        defaults={'libelle': libelle_sec, 'actif': True}
    )
    sections[code_sd] = section
    log(f"{'Créée' if created else 'Existe'} : {code_sec} — {libelle_sec}")


# ── Lecture Excel ─────────────────────────────────────────────────────────────
head("Lecture du fichier Excel")
FICHIER = os.path.join(BASE_DIR, 'Liste_collaborateurs.xlsx')
if not os.path.isfile(FICHIER):
    print(f"ERREUR : Fichier introuvable : {FICHIER}")
    sys.exit(1)

wb = openpyxl.load_workbook(FICHIER, data_only=True)
ws = wb['BASE DE DONNEES ']
log(f"{ws.max_row - 1} collaborateurs à traiter")


# ── Création des utilisateurs ─────────────────────────────────────────────────
head("Création des utilisateurs")
nb_crees = nb_ignores = nb_erreurs = 0

for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    if not row[0]:
        continue

    matricule = nettoyer(row[1])
    nom       = nettoyer(row[2])
    prenom    = nettoyer(row[3])
    poste     = nettoyer(row[4])
    genre     = nettoyer(row[5])
    sd_code   = nettoyer(row[8])
    categorie = nettoyer(row[9])

    if not nom:
        nb_ignores += 1
        continue

    # SD et section
    sd      = sds.get(sd_code)
    section = sections.get(sd_code)
    if not sd:
        warn(f"SD inconnue '{sd_code}' pour {nom} {prenom} — ignoré")
        nb_erreurs += 1
        continue

    # Rôle
    role_code = CAT_ROLE.get(categorie, 'visiteur')
    role      = roles.get(role_code)
    if not role:
        role = roles.get('visiteur')

    # Vérifier si le matricule existe déjà
    if matricule and Utilisateur.objects.filter(matricule=matricule).exists():
        nb_ignores += 1
        continue

    # Générer username
    username = generer_username(prenom, nom, idx)

    # Email
    email = f"{username}@cie.ci"

    # Créer l'utilisateur
    try:
        u = Utilisateur(
            username     = username,
            first_name   = prenom.title()[:150] if prenom else '',
            last_name    = nom.title()[:150]     if nom    else '',
            email        = email,
            matricule    = matricule if matricule else f"CIE-{idx:04d}",
            role         = role,
            est_actif_cie= True,
            is_active    = True,
        )
        u.set_password(MOT_DE_PASSE)
        u.save()

        # Affecter à la section
        if section:
            UtilisateurSection.objects.get_or_create(
                utilisateur=u,
                section=section,
                defaults={'est_principale': True}
            )

        nb_crees += 1
        log(f"{username:30} | {role_code:10} | {sd_code}")

    except Exception as e:
        warn(f"Erreur pour {nom} {prenom} : {e}")
        nb_erreurs += 1


# ── Résumé ────────────────────────────────────────────────────────────────────
print(f"""
{'='*60}
  IMPORT TERMINÉ
{'='*60}
  Utilisateurs créés  : {nb_crees}
  Déjà existants      : {nb_ignores}
  Erreurs             : {nb_erreurs}
  Mot de passe par défaut : {MOT_DE_PASSE}
{'='*60}
""")
