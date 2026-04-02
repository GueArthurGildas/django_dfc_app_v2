"""
Script d'import des activités depuis Classeur4.xlsx

Ce script :
  1. Crée les 12 sous-directions du fichier Excel dans l'application
  2. Crée une section pour chaque sous-direction
  3. Crée les dossiers rattachés à la bonne SD
  4. Crée les 166 activités avec leurs dates

Placer ce fichier ET Classeur4.xlsx dans le dossier app\

Usage :
    python seed_activites.py
"""
import os, sys
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django
django.setup()

import openpyxl
from datetime import date, datetime
from apps.organisation.models import SousDirection, Section
from apps.authentication.models import Utilisateur
from apps.operations.models import Dossier, Activite

def log(msg):  print(f"  ✓  {msg}")
def warn(msg): print(f"  ⚠  {msg}")
def head(msg): print(f"\n{'─'*60}\n  {msg}\n{'─'*60}")

# ── Couleurs disponibles dans l'app (COULEURS_SD) ─────────────────────────────
# '#003F7F','#0056A8','#1a7cc1','#17a2b8','#28a745',
# '#20c997','#6f42c1','#e83e8c','#fd7e14','#F5A623','#dc3545','#6c757d'

# ── Définition des 12 SD du fichier Excel ─────────────────────────────────────
SD_DEFINITIONS = [
    # (code,       libelle,                                          couleur,   ordre)
    ('SDCC',    'Sous-Direction Clientèle et Commerciale',           '#003F7F',  1),
    ('SDPCC',   'Sous-Direction Projet et Comptabilité',             '#0056A8',  2),
    ('SDCGR',   'Sous-Direction Comptabilité Générale',              '#6f42c1',  3),
    ('SDCF',    'Sous-Direction Comptabilité Financière',            '#dc3545',  4),
    ('TRESO',   'Sous-Direction Trésorerie',                         '#17a2b8',  5),
    ('DBCG',    'Direction Budget et Contrôle de Gestion',           '#fd7e14',  6),
    ('DFC-DBCG','DFC — Budget et Contrôle de Gestion',              '#F5A623',  7),
    ('SDF',     'Sous-Direction Fiscalité',                          '#1a7cc1',  8),
    ('SDSCS',   'Sous-Direction Stocks et Charges Sociales',         '#28a745',  9),
    ('DCEGPS',  'Direction Contrôle et Gestion des Projets',         '#20c997', 10),
    ('BMA',     'Bureau des Méthodes et Audit',                      '#6c757d', 11),
    ('Tous',    'Activités Transversales DFC',                       '#e83e8c', 12),
]

# ── Section principale par SD ─────────────────────────────────────────────────
SD_SECTIONS = {
    'SDCC':    ('SC-COM',    'Section Commerciale'),
    'SDPCC':   ('SC-COMPTA', 'Section Comptabilité'),
    'SDCGR':   ('SC-CG',     'Section Comptabilité Générale'),
    'SDCF':    ('SC-CF',     'Section Comptabilité Financière'),
    'TRESO':   ('SC-TRES',   'Section Trésorerie'),
    'DBCG':    ('SC-BCG',    'Section Budget et Contrôle'),
    'DFC-DBCG':('SC-DFCB',   'Section DFC-Budget'),
    'SDF':     ('SC-FISC',   'Section Fiscalité'),
    'SDSCS':   ('SC-STO',    'Section Stocks'),
    'DCEGPS':  ('SC-PROJ',   'Section Projets'),
    'BMA':     ('SC-AUD',    'Section Audit'),
    'Tous':    ('SC-TRANS',  'Section Transversale'),
}


# ── Utilisateur admin ─────────────────────────────────────────────────────────
head("Chargement des données en base")

admin = (Utilisateur.objects.filter(is_superuser=True).first()
         or Utilisateur.objects.filter(role__code='dfc').first())
if not admin:
    print("ERREUR : Aucun utilisateur admin. Lancez d'abord init_projet.py")
    sys.exit(1)
log(f"Utilisateur créateur : {admin.username}")


# ── Création des sous-directions ──────────────────────────────────────────────
head("Création des sous-directions")

sds = {}
for code, libelle, couleur, ordre in SD_DEFINITIONS:
    sd, created = SousDirection.objects.get_or_create(
        code=code,
        defaults={
            'libelle':      libelle,
            'couleur':      couleur,
            'actif':        True,
            'ordre':        ordre,
            'description':  f"Sous-direction {libelle}",
        }
    )
    sds[code] = sd
    log(f"{'Créée' if created else 'Existe'} : {code} — {libelle}")


# ── Création des sections ─────────────────────────────────────────────────────
head("Création des sections")

sections = {}
for code_sd, (code_sec, libelle_sec) in SD_SECTIONS.items():
    sd = sds[code_sd]
    section, created = Section.objects.get_or_create(
        code=code_sec,
        sous_direction=sd,
        defaults={'libelle': libelle_sec, 'actif': True}
    )
    sections[code_sd] = section
    log(f"{'Créée' if created else 'Existe'} : {code_sec} — {libelle_sec} ({code_sd})")


# ── Lecture Excel ─────────────────────────────────────────────────────────────
head("Lecture du fichier Excel")

FICHIER = os.path.join(BASE_DIR, 'Classeur4.xlsx')
if not os.path.isfile(FICHIER):
    print(f"ERREUR : {FICHIER} introuvable")
    print("Placez Classeur4.xlsx dans le même dossier que ce script.")
    sys.exit(1)

wb = openpyxl.load_workbook(FICHIER, data_only=True)
ws = wb.active
log(f"{ws.max_row - 1} activités à importer")


# ── Import des dossiers et activités ─────────────────────────────────────────
head("Import des dossiers et activités")

dossiers_cache = {}
nb_dossiers = nb_activites = nb_ignores = 0

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    return date.today()

for row in ws.iter_rows(min_row=2, values_only=True):
    titre_raw, debut, fin, dossier_nom, _, sd_code = row

    if not titre_raw or not str(titre_raw).strip():
        nb_ignores += 1
        continue

    titre       = str(titre_raw).strip()
    dossier_nom = str(dossier_nom).strip() if dossier_nom else 'Divers'
    sd_code     = str(sd_code).strip()     if sd_code     else 'Tous'

    sd      = sds.get(sd_code)
    section = sections.get(sd_code)

    if not sd or not section:
        warn(f"SD '{sd_code}' inconnue — ligne ignorée")
        nb_ignores += 1
        continue

    # Dossier — une clé par (SD + nom dossier)
    cle = f"{sd_code}|{dossier_nom}"
    if cle not in dossiers_cache:
        dossier_obj, created = Dossier.objects.get_or_create(
            titre=dossier_nom,
            section=section,
            defaults={
                'description': f"{dossier_nom} — {sd.libelle}",
                'responsable': admin,
                'est_actif':   True,
            }
        )
        dossiers_cache[cle] = dossier_obj
        if created:
            nb_dossiers += 1
            log(f"Dossier : [{sd_code}] {dossier_nom}")

    dossier_obj = dossiers_cache[cle]

    # Dates
    d_ouv = to_date(debut)
    d_but = to_date(fin)
    if d_but < d_ouv:
        d_but = d_ouv

    # Activité
    if not Activite.objects.filter(titre=titre, dossier=dossier_obj).exists():
        Activite.objects.create(
            titre           = titre,
            dossier         = dossier_obj,
            section         = section,
            type_activite   = 'ponctuelle',
            statut          = 'ouverte',
            date_ouverture  = d_ouv,
            date_butoir     = d_but,
            etat_avancement = 0,
            mois_reference  = d_ouv.strftime('%Y-%m'),
            created_by      = admin,
        )
        nb_activites += 1
    else:
        nb_ignores += 1


print(f"""
{'='*60}
  IMPORT TERMINÉ
{'='*60}
  Sous-directions : {len(sds)}
  Sections        : {len(sections)}
  Dossiers créés  : {nb_dossiers}
  Activités créées: {nb_activites}
  Ignorées        : {nb_ignores}
{'='*60}
""")
