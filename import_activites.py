"""
Script d'import des activités depuis Classeur4.xlsx
Utilise get_or_create — ne plante jamais sur les doublons

Placer ce fichier ET Classeur4.xlsx dans le dossier app\
Usage : python import_activites.py
"""
import os, sys
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django
django.setup()

import openpyxl
from datetime import date, datetime
from apps.organisation.models import SousDirection, Section
from apps.authentication.models import Utilisateur
from apps.operations.models import Dossier, Activite

def log(msg):  print(f"  ✓  {msg}")
def warn(msg): print(f"  ⚠  {msg}")
def head(msg): print(f"\n{'─'*60}\n  {msg}\n{'─'*60}")

# ── Définitions des SD ────────────────────────────────────────────────────────
SD_DEFINITIONS = [
    ('SDCC',     'Sous-Direction Clientèle et Commerciale',     '#003F7F',  1),
    ('SDPCC',    'Sous-Direction Projet et Comptabilité',       '#0056A8',  2),
    ('SDCGR',    'Sous-Direction Comptabilité Générale',        '#6f42c1',  3),
    ('SDCF',     'Sous-Direction Comptabilité Financière',      '#dc3545',  4),
    ('TRESO',    'Sous-Direction Trésorerie',                   '#17a2b8',  5),
    ('DBCG',     'Direction Budget et Contrôle de Gestion',     '#fd7e14',  6),
    ('DFC-DBCG', 'DFC — Budget et Contrôle de Gestion',        '#F5A623',  7),
    ('SDF',      'Sous-Direction Fiscalité',                    '#1a7cc1',  8),
    ('SDSCS',    'Sous-Direction Stocks et Charges Sociales',   '#28a745',  9),
    ('DCEGPS',   'Direction Contrôle et Gestion des Projets',   '#20c997', 10),
    ('BMA',      'Bureau des Méthodes et Audit',                '#6c757d', 11),
    ('Tous',     'Activités Transversales DFC',                 '#e83e8c', 12),
]

SD_SECTIONS = {
    'SDCC':     ('SC-COM',   'Section Commerciale'),
    'SDPCC':    ('SC-COMP',  'Section Comptabilité'),
    'SDCGR':    ('SC-CG',    'Section Comptabilité Générale'),
    'SDCF':     ('SC-CF',    'Section Comptabilité Financière'),
    'TRESO':    ('SC-TRES',  'Section Trésorerie'),
    'DBCG':     ('SC-BCG',   'Section Budget et Contrôle'),
    'DFC-DBCG': ('SC-DFCB',  'Section DFC-Budget'),
    'SDF':      ('SC-FISC',  'Section Fiscalité'),
    'SDSCS':    ('SC-STO',   'Section Stocks'),
    'DCEGPS':   ('SC-PROJ',  'Section Projets'),
    'BMA':      ('SC-AUD',   'Section Audit'),
    'Tous':     ('SC-TRANS', 'Section Transversale'),
}

# ── Utilisateur admin ─────────────────────────────────────────────────────────
head("Chargement utilisateur")
admin = (Utilisateur.objects.filter(is_superuser=True).first()
      or Utilisateur.objects.filter(role__code='dfc').first())
if not admin:
    print("ERREUR : Aucun utilisateur admin. Lancez d'abord init_projet.py")
    sys.exit(1)
log(f"Créateur : {admin.username}")

# ── Sous-directions ───────────────────────────────────────────────────────────
head("Sous-directions")
sds = {}
for code, libelle, couleur, ordre in SD_DEFINITIONS:
    sd, created = SousDirection.objects.get_or_create(
        code=code,
        defaults={
            'libelle':     libelle,
            'description': f"Sous-direction {libelle}",
            'mission':     '',
            'couleur':     couleur,
            'actif':       True,
            'ordre':       ordre,
        }
    )
    sds[code] = sd
    log(f"{'Créée' if created else 'Existe'} : {code} — {libelle}")

# ── Sections ──────────────────────────────────────────────────────────────────
head("Sections")
sections = {}
for code_sd, (code_sec, libelle_sec) in SD_SECTIONS.items():
    sd = sds[code_sd]
    section, created = Section.objects.get_or_create(
        code=code_sec,
        sous_direction=sd,
        defaults={'libelle': libelle_sec, 'actif': True}
    )
    sections[code_sd] = section
    log(f"{'Créée' if created else 'Existe'} : {code_sec} ({code_sd})")

# ── Lecture Excel ─────────────────────────────────────────────────────────────
head("Lecture Excel")
FICHIER = os.path.join(BASE_DIR, 'Classeur4.xlsx')
if not os.path.isfile(FICHIER):
    print(f"ERREUR : {FICHIER} introuvable")
    sys.exit(1)
wb = openpyxl.load_workbook(FICHIER, data_only=True)
ws = wb.active
log(f"{ws.max_row - 1} lignes à traiter")

# ── Dossiers et activités ─────────────────────────────────────────────────────
head("Dossiers et activités")
dossiers_cache = {}
nb_dos = nb_act = nb_skip = 0

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    return date.today()

for row in ws.iter_rows(min_row=2, values_only=True):
    titre_raw, debut, fin, dossier_nom, _, sd_code = row

    if not titre_raw or not str(titre_raw).strip():
        continue

    titre       = str(titre_raw).strip()
    dossier_nom = str(dossier_nom).strip() if dossier_nom else 'Divers'
    sd_code     = str(sd_code).strip()     if sd_code     else 'Tous'

    sd      = sds.get(sd_code)
    section = sections.get(sd_code)
    if not sd or not section:
        warn(f"SD inconnue : {sd_code}")
        nb_skip += 1
        continue

    # Dossier
    cle = f"{sd_code}|{dossier_nom}"
    if cle not in dossiers_cache:
        dossier_obj, created = Dossier.objects.get_or_create(
            titre=dossier_nom,
            section=section,
            defaults={
                'description': f"{dossier_nom} — {sd.libelle}",
                'responsable': admin,
                'est_actif':   True,
            }
        )
        dossiers_cache[cle] = dossier_obj
        if created:
            nb_dos += 1

    dossier_obj = dossiers_cache[cle]

    # Dates
    d_ouv = to_date(debut)
    d_but = to_date(fin)
    if d_but < d_ouv:
        d_but = d_ouv

    # Activité
    if not Activite.objects.filter(titre=titre, dossier=dossier_obj).exists():
        Activite.objects.create(
            titre           = titre,
            dossier         = dossier_obj,
            section         = section,
            type_activite   = 'ponctuelle',
            statut          = 'ouverte',
            date_ouverture  = d_ouv,
            date_butoir     = d_but,
            etat_avancement = 0,
            mois_reference  = d_ouv.strftime('%Y-%m'),
            created_by      = admin,
        )
        nb_act += 1
    else:
        nb_skip += 1

print(f"""
{'='*60}
  IMPORT TERMINÉ
{'='*60}
  Sous-directions : {len(sds)}
  Sections        : {len(sections)}
  Dossiers créés  : {nb_dos}
  Activités créées: {nb_act}
  Ignorées        : {nb_skip}
{'='*60}
""")
