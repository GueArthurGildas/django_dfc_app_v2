import json
import os
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Q

from .models import BgImport, BgLigne, BgAgregat, BgAlerte, BgAuditTrail, BgRegleAlerte
from .forms  import BgUploadForm, AlerteJustificationForm, ValidationForm

ROLES_BG = ('admin', 'dfc', 'da', 'sd', 'chef', 'cadre')


def _check_access(user):
    return user.role and user.role.code in ROLES_BG


# ── Dashboard ─────────────────────────────────────────────────────────────────

class BgDashboardView(LoginRequiredMixin, View):
    template_name = 'balance_generale/dashboard.html'

    def get(self, request):
        if not _check_access(request.user):
            messages.error(request, "Accès non autorisé.")
            return redirect('dashboard:index')

        exercice = request.GET.get('exercice', '')
        periode  = request.GET.get('periode', '')

        # Dernier import validé par défaut
        qs = BgImport.objects.all()
        if exercice:
            qs = qs.filter(exercice=exercice)
        if periode:
            qs = qs.filter(periode=periode)

        bg_actif   = qs.filter(statut__in=('valide','gele','traite')).first()
        imports    = BgImport.objects.order_by('-exercice', '-periode')[:24]
        exercices  = BgImport.objects.values_list('exercice', flat=True).distinct().order_by('-exercice')

        MOIS = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
        periodes = [(str(i), f"P{i:02d} — {MOIS[i-1]}") for i in range(1, 13)]

        context = {
            'bg_actif':  bg_actif,
            'imports':   imports,
            'exercices': exercices,
            'periodes':  periodes,
            'filtre_exercice': exercice,
            'filtre_periode':  periode,
        }

        if bg_actif:
            context.update(_build_dashboard_data(bg_actif))

        return render(request, self.template_name, context)


def _build_dashboard_data(bg):
    """Construit les données du dashboard pour un import donné."""
    noms_classes = {
        1:'Ressources durables', 2:'Actif immobilisé', 3:'Stocks',
        4:'Tiers', 5:'Trésorerie', 6:'Charges', 7:'Produits', 8:'Spéciaux'
    }

    # Agrégats par classe sous forme de liste structurée pour le template
    agregats_raw = BgAgregat.objects.filter(bg_import=bg, type_agregat='classe')
    agregats_dict = {f"{a.entite}_{a.classe}": a for a in agregats_raw}

    # Table des agrégats par classe : liste de dicts utilisable directement
    table_agregats = []
    for cl in range(1, 9):
        cie  = agregats_dict.get(f"CIE_{cl}")
        sect = agregats_dict.get(f"Secteur_{cl}")
        expl = agregats_dict.get(f"Exploitant_{cl}")
        if expl:
            table_agregats.append({
                'classe':       cl,
                'libelle':      noms_classes.get(cl, f'Classe {cl}'),
                'cie_total':    cie.valeur_total  if cie  else 0,
                'sect_total':   sect.valeur_total if sect else 0,
                'expl_total':   expl.valeur_total,
                'expl_report':  expl.valeur_report,
                'expl_periode': expl.valeur_periode,
                'expl_exercice':expl.valeur_exercice,
            })

    kpis_raw = BgAgregat.objects.filter(bg_import=bg, type_agregat='kpi', entite='Exploitant')
    kpis = [(k.libelle, k.valeur_total) for k in kpis_raw]

    alertes_critiques = bg.alertes.filter(niveau='critique', statut='ouverte').count()
    alertes_hautes    = bg.alertes.filter(niveau='haute',    statut='ouverte').count()
    alertes_moyennes  = bg.alertes.filter(niveau='moyenne',  statut='ouverte').count()

    chart_labels = [f"Cl.{r['classe']} {r['libelle']}" for r in table_agregats if r['expl_total'] != 0]
    chart_values = [abs(r['expl_total']) for r in table_agregats if r['expl_total'] != 0]

    return {
        'kpis':              kpis,
        'table_agregats':    table_agregats,
        'alertes_critiques': alertes_critiques,
        'alertes_hautes':    alertes_hautes,
        'alertes_moyennes':  alertes_moyennes,
        'chart_labels':      json.dumps(chart_labels),
        'chart_values':      json.dumps(chart_values),
    }


# ── Import ────────────────────────────────────────────────────────────────────

class BgImportView(LoginRequiredMixin, View):
    template_name = 'balance_generale/import.html'

    def get(self, request):
        if not _check_access(request.user):
            return redirect('dashboard:index')
        return render(request, self.template_name, {'form': BgUploadForm()})

    def post(self, request):
        if not _check_access(request.user):
            return redirect('dashboard:index')

        form = BgUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        fichier   = request.FILES['fichier']
        exercice  = form.cleaned_data['exercice']
        periode   = form.cleaned_data['periode']

        # Sauvegarder le fichier
        upload_dir = os.path.join('media', 'balance_generale')
        os.makedirs(upload_dir, exist_ok=True)
        nom_fichier = f"BG_{exercice}_P{periode:02d}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        chemin      = os.path.join(upload_dir, nom_fichier)
        with open(chemin, 'wb') as f:
            for chunk in fichier.chunks():
                f.write(chunk)

        # Créer l'import
        bg = BgImport.objects.create(
            fichier_nom  = fichier.name,
            fichier_path = chemin,
            exercice     = exercice,
            periode      = periode,
            importe_par  = request.user,
            statut       = 'traitement',
        )

        BgAuditTrail.objects.create(bg_import=bg, action='import', utilisateur=request.user,
            detail=f"Import du fichier {fichier.name}")

        # Traitement synchrone
        from .scripts.bg_processor import traiter_balance
        nb_lignes, nb_anomalies, erreur = traiter_balance(bg, chemin)

        if erreur:
            bg.statut = 'rejete'
            bg.erreur_traitement = erreur
            bg.save()
            messages.error(request, f"Erreur de traitement : {erreur[:200]}")
        else:
            bg.nb_lignes    = nb_lignes
            bg.nb_anomalies = nb_anomalies
            bg.statut       = 'traite'
            bg.save()
            messages.success(request,
                f"Balance Générale {bg.label_periode} importée : {nb_lignes} comptes, {nb_anomalies} anomalie(s) détectée(s).")

        return redirect('bg:detail', pk=bg.pk)


# ── Liste des imports ─────────────────────────────────────────────────────────

class BgListeImportsView(LoginRequiredMixin, View):
    template_name = 'balance_generale/liste_imports.html'

    def get(self, request):
        if not _check_access(request.user):
            return redirect('dashboard:index')
        imports = BgImport.objects.all().select_related('importe_par', 'valide_par')
        return render(request, self.template_name, {'imports': imports})


# ── Détail d'un import ────────────────────────────────────────────────────────

class BgDetailView(LoginRequiredMixin, View):
    template_name = 'balance_generale/detail.html'

    def get(self, request, pk):
        bg       = get_object_or_404(BgImport, pk=pk)
        alertes  = bg.alertes.all().order_by('niveau', 'num_compte')
        agregats = BgAgregat.objects.filter(bg_import=bg, type_agregat='classe').order_by('entite','classe')
        audit    = bg.audit.all()[:20]
        filtre_niveau = request.GET.get('niveau', '')
        if filtre_niveau:
            alertes = alertes.filter(niveau=filtre_niveau)

        context = {
            'bg':           bg,
            'alertes':      alertes,
            'agregats':     agregats,
            'audit':        audit,
            'form_valid':   ValidationForm(),
            'filtre_niveau':filtre_niveau,
            'stats_alertes':{
                'critique':          bg.alertes.filter(niveau='critique').count(),
                'haute':             bg.alertes.filter(niveau='haute').count(),
                'moyenne':           bg.alertes.filter(niveau='moyenne').count(),
                'ouvertes':          bg.alertes.filter(statut='ouverte').count(),
                'critique_ouvertes': bg.alertes.filter(niveau='critique', statut='ouverte').count(),
            }
        }
        context.update(_build_dashboard_data(bg))
        return render(request, self.template_name, context)


# ── Validation DF ─────────────────────────────────────────────────────────────

class BgValiderView(LoginRequiredMixin, View):
    def post(self, request, pk):
        bg   = get_object_or_404(BgImport, pk=pk)
        role = request.user.role.code if request.user.role else ''

        if role not in ('admin', 'dfc', 'da'):
            messages.error(request, "Seule la Direction Financière peut valider.")
            return redirect('bg:detail', pk=pk)

        if bg.alertes.filter(niveau='critique', statut='ouverte').exists():
            messages.error(request, "Des alertes critiques sont encore ouvertes. Traitez-les avant de valider.")
            return redirect('bg:detail', pk=pk)

        form = ValidationForm(request.POST)
        if form.is_valid():
            bg.statut          = 'gele' if form.cleaned_data.get('geler_periode') else 'valide'
            bg.valide_par      = request.user
            bg.date_validation = timezone.now()
            bg.commentaire_df  = form.cleaned_data.get('commentaire', '')
            bg.save()
            BgAuditTrail.objects.create(
                bg_import=bg, action='validation', utilisateur=request.user,
                detail=f"Validé {'et gelé ' if bg.statut=='gele' else ''}par {request.user.nom_complet}"
            )
            messages.success(request, f"Balance Générale {bg.label_periode} validée.")
        return redirect('bg:detail', pk=pk)


# ── Traitement alerte ─────────────────────────────────────────────────────────

class BgTraiterAlerteView(LoginRequiredMixin, View):
    def post(self, request, pk, alerte_pk):
        alerte = get_object_or_404(BgAlerte, pk=alerte_pk, bg_import_id=pk)
        form   = AlerteJustificationForm(request.POST, instance=alerte)
        if form.is_valid():
            a = form.save(commit=False)
            a.traite_par = request.user
            a.traite_le  = timezone.now()
            a.save()
            # Mettre à jour le compteur d'anomalies
            bg = alerte.bg_import
            bg.nb_anomalies = bg.alertes.filter(statut='ouverte').count()
            bg.save(update_fields=['nb_anomalies'])
            messages.success(request, "Alerte mise à jour.")
        return redirect('bg:detail', pk=pk)


# ── Vue comptes ───────────────────────────────────────────────────────────────

class BgComptesView(LoginRequiredMixin, View):
    template_name = 'balance_generale/comptes.html'

    def get(self, request, pk):
        bg       = get_object_or_404(BgImport, pk=pk)
        entite   = request.GET.get('entite', 'Exploitant')
        classe   = request.GET.get('classe', '')
        q        = request.GET.get('q', '').strip()

        lignes = BgLigne.objects.filter(bg_import=bg, entite=entite)
        if classe:
            lignes = lignes.filter(classe=classe)
        if q:
            lignes = lignes.filter(Q(num_compte__icontains=q) | Q(intitule__icontains=q))
        lignes = lignes.order_by('num_compte')

        return render(request, self.template_name, {
            'bg':      bg,
            'lignes':  lignes,
            'entite':  entite,
            'classe':  classe,
            'q':       q,
        })


# ── API JSON pour Chart.js ────────────────────────────────────────────────────

class BgApiKpisView(LoginRequiredMixin, View):
    def get(self, request, pk):
        bg   = get_object_or_404(BgImport, pk=pk)
        data = _build_dashboard_data(bg)
        return JsonResponse({
            'kpis':          data.get('kpis', {}),
            'chart_labels':  json.loads(data.get('chart_labels', '[]')),
            'chart_values':  json.loads(data.get('chart_values', '[]')),
            'alertes': {
                'critique': data.get('alertes_critiques', 0),
                'haute':    data.get('alertes_hautes', 0),
                'moyenne':  data.get('alertes_moyennes', 0),
            }
        })


# ── Suppression d'un import ──────────────────────────────────────────────────

class BgSupprimerView(LoginRequiredMixin, View):
    def post(self, request, pk):
        bg   = get_object_or_404(BgImport, pk=pk)
        role = request.user.role.code if request.user.role else ''

        if role not in ('admin', 'dfc', 'da'):
            messages.error(request, "Seuls Admin, DFC et DA peuvent supprimer un import.")
            return redirect('bg:detail', pk=pk)

        if bg.statut == 'gele':
            messages.error(request, "Impossible de supprimer une période gelée.")
            return redirect('bg:detail', pk=pk)

        label = bg.label_periode
        # Supprimer le fichier Excel si présent
        import os
        if bg.fichier_path and os.path.isfile(bg.fichier_path):
            try:
                os.remove(bg.fichier_path)
            except Exception:
                pass
        # Supprimer l'import (cascade → lignes, agrégats, alertes, audit)
        bg.delete()
        messages.success(request, f"Balance Générale {label} supprimée.")
        return redirect('bg:imports')


# ── Export Excel ──────────────────────────────────────────────────────────────

class BgExportExcelView(LoginRequiredMixin, View):
    def get(self, request, pk):
        bg = get_object_or_404(BgImport, pk=pk)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messages.error(request, "openpyxl non installé. Lancez : pip install openpyxl")
            return redirect('bg:detail', pk=pk)

        wb = openpyxl.Workbook()

        # Feuille 1 : Données par classe
        ws1 = wb.active
        ws1.title = f"BG {bg.label_periode}"
        entetes = ['Entité', 'Classe', 'Report', 'Activité Période', 'Activité Exercice', 'Total']
        for col, titre in enumerate(entetes, 1):
            c = ws1.cell(row=1, column=col, value=titre)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='003F7F')

        row = 2
        for a in BgAgregat.objects.filter(bg_import=bg, type_agregat='classe').order_by('entite','classe'):
            ws1.cell(row=row, column=1, value=a.entite)
            ws1.cell(row=row, column=2, value=f"Classe {a.classe}")
            ws1.cell(row=row, column=3, value=a.valeur_report)
            ws1.cell(row=row, column=4, value=a.valeur_periode)
            ws1.cell(row=row, column=5, value=a.valeur_exercice)
            ws1.cell(row=row, column=6, value=a.valeur_total)
            row += 1

        # Feuille 2 : Alertes
        ws2 = wb.create_sheet("Alertes")
        for col, titre in enumerate(['Niveau','Code','Compte','Entité','Message','Statut'], 1):
            c = ws2.cell(row=1, column=col, value=titre)
            c.font = Font(bold=True)

        for i, a in enumerate(BgAlerte.objects.filter(bg_import=bg).order_by('niveau'), 2):
            ws2.cell(row=i, column=1, value=a.niveau)
            ws2.cell(row=i, column=2, value=a.code_regle)
            ws2.cell(row=i, column=3, value=a.num_compte)
            ws2.cell(row=i, column=4, value=a.entite)
            ws2.cell(row=i, column=5, value=a.message)
            ws2.cell(row=i, column=6, value=a.statut)

        BgAuditTrail.objects.create(bg_import=bg, action='export', utilisateur=request.user,
            detail="Export Excel")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="BG_{bg.exercice}_P{bg.periode:02d}.xlsx"'
        wb.save(response)
        return response
